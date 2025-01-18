import os
import pandas as pd
from openpyxl import load_workbook

def __load_yaml(file_path):
    with open(file_path, "r") as file:
        return file.read()

def load_predictions(
        type: str,
        file_path:str="/Users/ejowik001/Desktop/Github/Nowcasting/kedro/refinery/data/08_reporting/dash_input_report.xlsx",):

    if os.path.exists(file_path):
        wb = load_workbook(file_path, read_only=True)
        if type == "base":
            assert "Nowcast Browser – Base" in wb.sheetnames
            df = pd.read_excel(file_path, sheet_name="Nowcast Browser – Base")
        elif type == "adj":
            assert "Nowcast Browser – Adjusted" in wb.sheetnames
            return None
            # TODO: Implement the adjusted data
            # df = pd.read_excel(file_path, sheet_name="Nowcast Browser – Adjusted")
        else: raise ValueError("Invalid type. Please choose 'base' or 'adj'.")

        df["dt"] = df["dt"].astype(str).replace(r"-\d{2}$", "", regex=True)

        return {
            "labels": [label if index % 3 == 0 else "" for index, label in enumerate(df["dt"])],
            "series": [df[c].tolist() for c in df.columns if c != "dt"]
        }
    return None

# Global variables to store ...
# pipeline_status = "Not started"
# last_run_timestamp = "N/A"
TARGET_FOLDER = "/Users/ejowik001/Desktop/Github/Nowcasting/kedro/refinery/data/_test"
PARAMETERS_PATH = "/Users/ejowik001/Desktop/Github/Nowcasting/kedro/refinery/conf/base/parameters.yml"
CATALOG_PATH= "/Users/ejowik001/Desktop/Github/Nowcasting/kedro/refinery/conf/base/catalog.yml"
parameters, data_catalog = __load_yaml(PARAMETERS_PATH), __load_yaml(CATALOG_PATH)

